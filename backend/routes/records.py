import io
import csv
from datetime import datetime
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId
from database import records, uploads
from services.validation_service import run_all_validations
import openpyxl
from openpyxl.styles import PatternFill, Font

router = APIRouter()

FIELD_NAMES = ["date", "shift", "employee_number", "operation_code",
               "machine_number", "work_order_number", "quantity_produced", "time_taken"]


def _serialize(doc: dict) -> dict:
    doc["id"] = str(doc.pop("_id"))
    for key in ["created_at", "review_time"]:
        if key in doc and doc[key] and hasattr(doc[key], "isoformat"):
            doc[key] = doc[key].isoformat()
    if "audit_trail" in doc:
        for entry in doc["audit_trail"]:
            if "timestamp" in entry and hasattr(entry["timestamp"], "isoformat"):
                entry["timestamp"] = entry["timestamp"].isoformat()
    return doc


def _get_field_val(record, field):
    f = record.get(field)
    if isinstance(f, dict):
        return f.get("value", "")
    return f or ""


@router.get("/records")
async def get_records(
    shift: str = Query(None),
    status: str = Query(None),
    date_from: str = Query(None),
    date_to: str = Query(None),
    search: str = Query(None),
    upload_id: str = Query(None),
    page: int = 1,
    limit: int = 20,
):
    query = {}
    if upload_id:
        query["upload_id"] = upload_id
    if shift:
        query["shift.value"] = shift.upper()
    if search:
        query["$or"] = [
            {"work_order_number.value": {"$regex": search, "$options": "i"}},
            {"machine_number.value": {"$regex": search, "$options": "i"}},
            {"employee_number.value": {"$regex": search, "$options": "i"}},
            {"operation_code.value": {"$regex": search, "$options": "i"}},
        ]
    if status:
        if status == "reviewed":
            query["reviewed"] = True
        elif status == "pending":
            query["reviewed"] = False

    skip = (page - 1) * limit
    sort_dir = 1 if upload_id else -1
    cursor = records.find(query).sort("created_at", sort_dir).skip(skip).limit(limit)
    docs = []
    async for doc in cursor:
        docs.append(_serialize(doc))
    total = await records.count_documents(query)
    return {"records": docs, "total": total, "page": page, "limit": limit}


@router.get("/records/export/csv")
async def export_csv():
    cursor = records.find({}).sort("created_at", -1)
    output = io.StringIO()
    writer = csv.writer(output)
    headers = ["ID", "Upload ID", "Date", "Shift", "Employee", "Operation Code",
               "Machine", "Work Order", "Qty Produced", "Time Taken",
               "Validation Errors", "Anomaly Flags", "Reviewed", "Created At"]
    writer.writerow(headers)
    async for doc in cursor:
        writer.writerow([
            str(doc.get("_id", "")),
            doc.get("upload_id", ""),
            _get_field_val(doc, "date"),
            _get_field_val(doc, "shift"),
            _get_field_val(doc, "employee_number"),
            _get_field_val(doc, "operation_code"),
            _get_field_val(doc, "machine_number"),
            _get_field_val(doc, "work_order_number"),
            _get_field_val(doc, "quantity_produced"),
            _get_field_val(doc, "time_taken"),
            "; ".join(doc.get("validation_errors", [])),
            "; ".join(doc.get("anomaly_flags", [])),
            doc.get("reviewed", False),
            doc.get("created_at", "").isoformat() if hasattr(doc.get("created_at", ""), "isoformat") else "",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=opscenter_records.csv"}
    )


@router.get("/records/export/excel")
async def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OpsCenter Records"
    headers = ["ID", "Upload ID", "Date", "Shift", "Employee", "Operation Code",
               "Machine", "Work Order", "Qty Produced", "Time Taken",
               "Validation Errors", "Anomaly Flags", "Reviewed", "Created At"]
    header_fill = PatternFill(start_color="00f5c4", end_color="00f5c4", fill_type="solid")
    header_font = Font(bold=True, color="020d12")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    green_fill = PatternFill(start_color="00ff88", end_color="00ff88", fill_type="solid")
    red_fill = PatternFill(start_color="ff4444", end_color="ff4444", fill_type="solid")

    cursor = records.find({}).sort("created_at", -1)
    row_num = 2
    async for doc in cursor:
        has_errors = len(doc.get("validation_errors", [])) > 0
        fill = red_fill if has_errors else green_fill
        row_data = [
            str(doc.get("_id", "")),
            doc.get("upload_id", ""),
            _get_field_val(doc, "date"),
            _get_field_val(doc, "shift"),
            _get_field_val(doc, "employee_number"),
            _get_field_val(doc, "operation_code"),
            _get_field_val(doc, "machine_number"),
            _get_field_val(doc, "work_order_number"),
            _get_field_val(doc, "quantity_produced"),
            _get_field_val(doc, "time_taken"),
            "; ".join(doc.get("validation_errors", [])),
            "; ".join(doc.get("anomaly_flags", [])),
            "YES" if doc.get("reviewed") else "NO",
            str(doc.get("created_at", "")),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
        row_num += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=opscenter_records.xlsx"}
    )


@router.get("/records/{record_id}")
async def get_record(record_id: str):
    try:
        doc = await records.find_one({"_id": ObjectId(record_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid record ID")
    if not doc:
        raise HTTPException(status_code=404, detail="Record not found")
    return _serialize(doc)


@router.put("/records/{record_id}")
async def update_record(record_id: str, body: dict):
    try:
        oid = ObjectId(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid record ID")

    existing = await records.find_one({"_id": oid})
    if not existing:
        raise HTTPException(status_code=404, detail="Record not found")

    audit_entries = []
    update_fields = {}

    for field in FIELD_NAMES:
        if field in body:
            new_val = body[field]
            original = existing.get(field, {})
            original_val = original.get("value") if isinstance(original, dict) else original
            if str(new_val) != str(original_val):
                audit_entries.append({
                    "field": field,
                    "original_value": original_val,
                    "corrected_value": new_val,
                    "timestamp": datetime.utcnow(),
                })
                update_fields[field] = {"value": new_val, "confidence": 1.0}

    # Re-run validation
    merged = dict(existing)
    merged.update(update_fields)
    py_errors = await run_all_validations(merged, exclude_id=record_id)
    error_messages = [f"{e['field']}: {e['message']}" for e in py_errors]

    update_fields["validation_errors"] = error_messages
    update_fields["reviewed"] = True
    update_fields["review_time"] = datetime.utcnow()

    if audit_entries:
        existing_trail = existing.get("audit_trail", [])
        existing_trail.extend(audit_entries)
        update_fields["audit_trail"] = existing_trail

    await records.update_one({"_id": oid}, {"$set": update_fields})

    # Update upload status
    upload_id = existing.get("upload_id")
    if upload_id:
        try:
            await uploads.update_one({"_id": ObjectId(upload_id)}, {"$set": {"status": "reviewed"}})
        except Exception:
            pass

    updated = await records.find_one({"_id": oid})
    return _serialize(updated)


@router.delete("/records/{record_id}")
async def delete_record(record_id: str):
    try:
        oid = ObjectId(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid record ID")

    record = await records.find_one({"_id": oid})
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")

    upload_id = record.get("upload_id")

    # Delete the record
    result = await records.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Failed to delete record")

    # If no records are left for this upload, update or delete the upload record
    if upload_id:
        remaining_count = await records.count_documents({"upload_id": upload_id})
        if remaining_count == 0:
            try:
                await uploads.delete_one({"_id": ObjectId(upload_id)})
            except Exception:
                pass

    return {"status": "success", "message": "Record deleted successfully"}
